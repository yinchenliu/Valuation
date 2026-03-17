"""Audit: compare raw Excel data vs parsed dataclass fields for GOOGL 2024."""
import sys
sys.path.insert(0, r"C:\Users\yinchenliu\Desktop\Python\python\Scripts\valuation_platform")

import openpyxl
from ingestion.capital_iq_parser import parse_capital_iq

FILE = r"C:\Users\yinchenliu\Downloads\AlphabetInc.NASDAQGSGOOGL_Report_03-02-2026.xlsx"
YEAR = 2024  # Check latest year

# --- Parse ---
financials = parse_capital_iq(FILE, "GOOGL", "Alphabet Inc.")
bs = financials.get_balance_sheet(YEAR)
inc = financials.get_income_statement(YEAR)
cf = financials.get_cash_flow(YEAR)

# --- Raw Excel data (2024 FY = column B, row 12 header) ---
wb = openpyxl.load_workbook(FILE, data_only=True)

def get_raw(sheet_name):
    """Return {label: value} for 2024 column."""
    ws = wb[sheet_name]
    data = {}
    for r in range(16, ws.max_row + 1):
        label = ws.cell(r, 1).value
        val = ws.cell(r, 2).value  # 2024 FY column
        if label and val is not None:
            data[str(label).strip()] = val
    return data

print("=" * 80)
print("BALANCE SHEET — 2024")
print("=" * 80)
raw_bs = get_raw("Balance Sheet (As Reported)")
for label, val in raw_bs.items():
    print(f"  Excel: {label:55s} = {val}")

print(f"\n  --- Parsed B/S fields ---")
print(f"  cash_and_equivalents:       {bs.cash_and_equivalents:>12,.0f}   (Excel: 23466)")
print(f"  short_term_investments:     {bs.short_term_investments:>12,.0f}   (Excel: 72191 Marketable Securities)")
print(f"  accounts_receivable:        {bs.accounts_receivable:>12,.0f}   (Excel: 52340)")
print(f"  inventory:                  {bs.inventory:>12,.0f}   (Excel: NA)")
print(f"  other_current_assets:       {bs.other_current_assets:>12,.0f}   (Excel: 15714)")
print(f"  total_current_assets:       {bs.total_current_assets:>12,.0f}   (Excel: 163711)")
print(f"  ppe_net:                    {bs.ppe_net:>12,.0f}   (Excel: 171036)")
print(f"  goodwill:                   {bs.goodwill:>12,.0f}   (Excel: 31885)")
print(f"  intangible_assets:          {bs.intangible_assets:>12,.0f}   (Excel: NA)")
print(f"  other_non_current_assets:   {bs.other_non_current_assets:>12,.0f}   (Excel: 14874)")
print(f"  total_assets:               {bs.total_assets:>12,.0f}   (Excel: 450256)")
print(f"  accounts_payable:           {bs.accounts_payable:>12,.0f}   (Excel: 7987)")
print(f"  short_term_debt:            {bs.short_term_debt:>12,.0f}   (Excel: N/A)")
print(f"  current_portion_lt_debt:    {bs.current_portion_lt_debt:>12,.0f}   (Excel: N/A)")
print(f"  accrued_liabilities:        {bs.accrued_liabilities:>12,.0f}   (Excel: 15069)")
print(f"  other_current_liabilities:  {bs.other_current_liabilities:>12,.0f}   (Excel: 51228)")
print(f"  total_current_liabilities:  {bs.total_current_liabilities:>12,.0f}   (Excel: 89122)")
print(f"  long_term_debt:             {bs.long_term_debt:>12,.0f}   (Excel: 10883)")
print(f"  other_non_current_liab:     {bs.other_non_current_liabilities:>12,.0f}   (Excel: 11691+8782+4694=25167)")
print(f"  total_liabilities:          {bs.total_liabilities:>12,.0f}   (Excel: 450256-325084=125172)")
print(f"  total_equity:               {bs.total_equity:>12,.0f}   (Excel: 325084)")
print(f"  total_debt:                 {bs.total_debt:>12,.0f}   (Excel: only LT 10883)")
print(f"  net_debt:                   {bs.net_debt:>12,.0f}")
print(f"  net_working_capital:        {bs.net_working_capital:>12,.0f}")

# Compute what's MISSING from B/S
missing_ca = 163711 - bs.total_current_assets
missing_ta = 450256 - bs.total_assets
missing_cl = 89122 - bs.total_current_liabilities
print(f"\n  GAPS:")
print(f"    Missing current assets:      {missing_ca:>10,.0f}  (Incomes Taxes Receivable=NA)")
print(f"    Missing total assets:        {missing_ta:>10,.0f}  (Oper Lease 13588 + Non-mkt Sec 37982 + Def Tax 17180)")
print(f"    Missing current liabilities: {missing_cl:>10,.0f}  (Deferred Rev 5036 + Accrued Rev Share 9802)")

print("\n" + "=" * 80)
print("INCOME STATEMENT — 2024")
print("=" * 80)
raw_is = get_raw("Income Statement (As Reported)")
for label, val in raw_is.items():
    print(f"  Excel: {label:55s} = {val}")

print(f"\n  --- Parsed I/S fields ---")
print(f"  revenue:                    {inc.revenue:>12,.0f}   (Excel: 350018)")
print(f"  cost_of_revenue:            {inc.cost_of_revenue:>12,.0f}   (Excel: -146306)")
print(f"  sga:                        {inc.sga:>12,.0f}   (Excel: S&M -27808 + G&A -14188)")
print(f"  rd_expense:                 {inc.rd_expense:>12,.0f}   (Excel: -49326)")
print(f"  depreciation_amortization:  {inc.depreciation_amortization:>12,.0f}   (Excel: N/A in I/S)")
print(f"  other_operating_expense:    {inc.other_operating_expense:>12,.0f}   (Excel: N/A)")
print(f"  interest_expense:           {inc.interest_expense:>12,.0f}   (Excel: N/A — bundled in Other Inc/Exp)")
print(f"  interest_income:            {inc.interest_income:>12,.0f}   (Excel: N/A)")
print(f"  other_non_operating:        {inc.other_non_operating:>12,.0f}   (Excel: 7425 Other Income/expense-net)")
print(f"  tax_expense:                {inc.tax_expense:>12,.0f}   (Excel: -19697)")
print(f"  diluted_shares_outstanding: {inc.diluted_shares_outstanding:>12,.0f}   (Excel: N/A — not in CIQ I/S)")
print(f"  --- Computed ---")
print(f"  gross_profit:               {inc.gross_profit:>12,.0f}")
print(f"  ebit:                       {inc.ebit:>12,.0f}   (Excel Operating Income: 112390)")
print(f"  ebt:                        {inc.ebt:>12,.0f}   (Excel EBT: 119815)")
print(f"  net_income:                 {inc.net_income:>12,.0f}   (Excel Net Income: 100118)")
print(f"  effective_tax_rate:         {inc.effective_tax_rate:>11.2%}   (19697/119815=16.4%)")
print(f"  operating_margin:           {inc.operating_margin:>11.2%}")

print("\n" + "=" * 80)
print("CASH FLOW — 2024")
print("=" * 80)
raw_cf = get_raw("Cash Flow (As Reported)")
for label, val in raw_cf.items():
    print(f"  Excel: {label:55s} = {val}")

print(f"\n  --- Parsed C/F fields ---")
print(f"  net_income:                 {cf.net_income:>12,.0f}   (Excel: 100118)")
print(f"  depreciation_amortization:  {cf.depreciation_amortization:>12,.0f}   (Excel: 15311 + NA)")
print(f"  stock_based_compensation:   {cf.stock_based_compensation:>12,.0f}   (Excel: 22785)")
print(f"  change_in_working_capital:  {cf.change_in_working_capital:>12,.0f}   (not in CIQ — derived from B/S)")
print(f"  other_operating_activities: {cf.other_operating_activities:>12,.0f}   (not captured)")
print(f"  capital_expenditures:       {cf.capital_expenditures:>12,.0f}   (Excel: -52535)")
print(f"  acquisitions:               {cf.acquisitions:>12,.0f}   (Excel: -2931)")
print(f"  other_investing_activities: {cf.other_investing_activities:>12,.0f}   (Excel: -2667)")
print(f"  debt_issued:                {cf.debt_issued:>12,.0f}   (Excel: 13589)")
print(f"  debt_repaid:                {cf.debt_repaid:>12,.0f}   (Excel: -12701)")
print(f"  shares_issued:              {cf.shares_issued:>12,.0f}   (Excel: N/A)")
print(f"  shares_repurchased:         {cf.shares_repurchased:>12,.0f}   (Excel: -62222)")
print(f"  dividends_paid:             {cf.dividends_paid:>12,.0f}   (Excel: -7363)")
print(f"  other_financing_activities: {cf.other_financing_activities:>12,.0f}")
print(f"  --- Computed ---")
print(f"  cash_from_operations:       {cf.cash_from_operations:>12,.0f}   (Excel CFO: 125299)")
print(f"  cash_from_investing:        {cf.cash_from_investing:>12,.0f}   (Excel CFI: -45536)")
print(f"  cash_from_financing:        {cf.cash_from_financing:>12,.0f}   (Excel CFF: -79733)")

# CFO gap
cfo_gap = 125299 - cf.cash_from_operations
cfi_gap = -45536 - cf.cash_from_investing
cff_gap = -79733 - cf.cash_from_financing
print(f"\n  GAPS:")
print(f"    CFO gap: {cfo_gap:>10,.0f}  (unmapped operating items)")
print(f"    CFI gap: {cfi_gap:>10,.0f}  (unmapped: Purch/Sales mkt sec, non-mkt sec)")
print(f"    CFF gap: {cff_gap:>10,.0f}  (unmapped: stock-based award payments, sale of interest)")
